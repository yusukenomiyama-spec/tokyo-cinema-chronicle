"""
scripts/build.py
data/movies.xlsx を読み込み、index.html のエントリ配列を更新する
Usage: python3 scripts/build.py

CSV 列:
  id, yearDisplay, yearSort, genre, icon, events, movieList, movieGenres
  - movieList / movieGenres は "|" 区切りで複数作品を記載
  - movieGenres は movieList が1作品のときは空欄でよい
"""

import csv
import json
import os

ROOT      = os.path.join(os.path.dirname(__file__), '..')
XLSX_PATH = os.path.join(ROOT, 'data', 'movies.xlsx')
CSV_PATH  = os.path.join(ROOT, 'data', 'movies.csv')
HTML_PATH = os.path.join(ROOT, 'index.html')

START_MARKER = '// ═══ MOVIE DATA START ═══'
END_MARKER   = '// ═══ MOVIE DATA END ═══'

# ── xlsx / CSV どちらからでも読み込む ──────────────────────────────
def load_from_xlsx():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb['映画データ']

    # 列は固定位置: A=id, B=yearDisplay, C=yearSort, D=genre,
    #               E=icon, F=events, G=movieList, H=movieGenres
    # データは 3行目から（1行目=タイトル, 2行目=ヘッダー）
    KEYS = ['id', 'yearDisplay', 'yearSort', 'genre', 'icon', 'events', 'movieList', 'movieGenres']

    rows = []
    for row in ws.iter_rows(min_row=3, max_col=8, values_only=True):
        raw_id = row[0]
        if raw_id is None or str(raw_id).strip() == '':
            continue
        try:
            int(float(str(raw_id)))
        except ValueError:
            continue

        def v(val):
            return str(val).strip() if val is not None else ''

        movie_list_val = v(row[6])
        if not movie_list_val:  # movieList が空の行（入力用空欄）はスキップ
            continue

        rows.append({
            'id':          v(row[0]),
            'yearDisplay': v(row[1]),
            'yearSort':    v(row[2]),
            'genre':       v(row[3]),
            'icon':        v(row[4]),
            'events':      v(row[5]),
            'movieList':   movie_list_val,
            'movieGenres': v(row[7]),
        })
    return rows

def load_from_csv():
    rows = []
    with open(CSV_PATH, encoding='utf-8', newline='') as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows

# xlsx を優先、なければ CSV にフォールバック
if os.path.exists(XLSX_PATH):
    try:
        raw_rows = load_from_xlsx()
        print(f'📖 xlsx から読み込み: {len(raw_rows)} 件')
    except Exception as e:
        print(f'⚠️  xlsx 読み込み失敗 ({e})、CSV にフォールバック')
        raw_rows = load_from_csv()
else:
    raw_rows = load_from_csv()
    print(f'📖 CSV から読み込み: {len(raw_rows)} 件')

# ── dict → Entry オブジェクト ─────────────────────────────────────
entries = []
for row in raw_rows:
    movie_list   = [m.strip() for m in str(row.get('movieList',   '')).split('|') if m.strip()]
    movie_genres = [g.strip() for g in str(row.get('movieGenres', '')).split('|') if g.strip()]

    year_sort_raw = str(row.get('yearSort', '0')).strip()
    try:
        year_sort = float(year_sort_raw)
    except ValueError:
        year_sort = 0.0

    entry = {
        'id':          int(float(str(row['id']))),
        'yearDisplay': str(row.get('yearDisplay', '')).strip(),
        'yearSort':    year_sort,
        'genre':       str(row.get('genre', '')).strip(),
        'icon':        str(row.get('icon',  '')).strip(),
        'events':      str(row.get('events','  ')).strip(),
        'movieList':   movie_list,
    }
    if len(movie_list) > 1 and movie_genres:
        entry['movieGenres'] = movie_genres

    entries.append(entry)

# id でソート
entries.sort(key=lambda e: e['id'])

# ── CSV を同期出力（xlsx 更新後に CSV も最新にしておく） ────────────
with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['id','yearDisplay','yearSort','genre','icon','events','movieList','movieGenres'])
    for e in entries:
        writer.writerow([
            e['id'], e['yearDisplay'], e['yearSort'],
            e['genre'], e['icon'], e['events'],
            '|'.join(e['movieList']),
            '|'.join(e.get('movieGenres', [])),
        ])

# ── Python dict → JS オブジェクト文字列 ──────────────────────────
def esc(s):
    return s.replace('\\', '\\\\').replace("'", "\\'")

def entry_to_js(e):
    movie_list_js   = json.dumps(e['movieList'],   ensure_ascii=False)
    movie_genres_js = (
        f", movieGenres:{json.dumps(e['movieGenres'], ensure_ascii=False)}"
        if 'movieGenres' in e else ''
    )
    year_sort = int(e['yearSort']) if e['yearSort'] == int(e['yearSort']) else e['yearSort']
    return (
        f"  {{ id:{e['id']}, yearDisplay:'{esc(e['yearDisplay'])}', "
        f"yearSort:{year_sort}, genre:'{esc(e['genre'])}', icon:'{esc(e['icon'])}', "
        f"events:'{esc(e['events'])}', movieList:{movie_list_js}{movie_genres_js} }}"
    )

entries_js = ',\n'.join(entry_to_js(e) for e in entries)
new_block = (
    f"{START_MARKER}\n"
    f"let entries = [\n{entries_js},\n];\n"
    f"{END_MARKER}"
)

# ── index.html を更新 ─────────────────────────────────────────────
with open(HTML_PATH, encoding='utf-8') as f:
    html = f.read()

start_idx = html.index(START_MARKER)
end_idx   = html.index(END_MARKER) + len(END_MARKER)

updated = html[:start_idx] + new_block + html[end_idx:]

with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(updated)

print(f'✅ ビルド完了: {len(entries)} 件のエントリを index.html に反映しました')
